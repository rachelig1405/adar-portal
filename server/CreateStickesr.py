import base64
import re
import barcode
from barcode.writer import ImageWriter
from pathlib import Path

from tkinter.ttk import Progressbar
import subprocess
import tempfile
import shutil
import pandas as pd
from jinja2 import Template
from playwright.sync_api import sync_playwright
from openpyxl import load_workbook
from io import BytesIO

from barcode import EAN13
from barcode.writer import SVGWriter


# =========================
# שמות עמודות באקסל
# =========================
COL_SKU = " ITEM  NUMBER"          # מק"ט / קוד פנימי
COL_ITEM_NO = "STICKER ITEM NO"    # ITEM NO
COL_PCS = "OUTER"                  # PCS
COL_CBM = "CBM"                    # CBM
COL_BARCODE = "BARCODE"            # ברקוד EAN13 - 12/13 ספרות
COL_PHOTO = "PHOTO"                # נתיב תמונת מוצר במחשב
COL_WARNING_TYPE = "סוג סטיקר"   # 0 רגיל, 1 רחפנים, 2 מכוניות
LOGO_PATH = Path(__file__).parent / "לוגו.png"
COL_PRODUCER="יצרן"
COL_AGE="גיל שימוש"
COL_X_STICKER="מידה X"
COL_Y_STICKER="מידה Y"
COL_BATTERY="בטריות"
COL_CERTIFICATE="מספר תעודה"
# העמודות שחייב להיות בהן ערך בכל שורת מוצר
REQUIRED_VALUE_COLUMNS = {
    COL_SKU: "מק״ט",
    COL_ITEM_NO: "ITEM NO",
    COL_PCS: "כמות בקרטון / OUTER",
    COL_CBM: "CBM",
    COL_BARCODE: "ברקוד",
   
    COL_PRODUCER: "יצרן",
    COL_AGE: "גיל שימוש",
}
# =========================================================
# מידות הברקוד בשלושת סוגי ה-PDF
# קובעים רק רוחב. הגובה מחושב אוטומטית בלי עיוות.
# =========================================================

BARCODE_WIDTH_MR_MM = 68.0
BARCODE_WIDTH_HTS_MM = 25.0
BARCODE_WIDTH_ST_MM = 31.0

# =========================================================
# שלוש תיקיות קבועות שבהן מחפשים תיקייה ישנה בשם המק"ט.
# החליפי את הנתיבים כאן בנתיבים האמיתיים אצלך.
# החיפוש מתבצע לפי הסדר: 1, אחר כך 2, אחר כך 3.
# =========================================================
OLD_PRODUCTS_ROOT_1 = Path(r"E:\Adar Shmuel 26 ltd Dropbox\tetroashvili shmuel\A תיק מוצר לכל המוצרים\0מוצרים ללא הזמנה")
OLD_PRODUCTS_ROOT_2 = Path(r"E:\OLD_PRODUCTS_2")
OLD_PRODUCTS_ROOT_3 = Path(r"E:\Adar Shmuel 26 ltd Dropbox\tetroashvili shmuel\A תיק מוצר לכל המוצרים\0מקטים שנבדקו")

OLD_PRODUCTS_ROOTS = [
    OLD_PRODUCTS_ROOT_1,
    OLD_PRODUCTS_ROOT_2,
    OLD_PRODUCTS_ROOT_3,
]
def validate_product_row(row, excel_row_number: int) -> list[str]:
    """
    בודקת שורת מוצר אחת ומחזירה רשימת שגיאות.
    אם הרשימה ריקה, השורה תקינה.
    """

    errors = []

    sku = val(row, COL_SKU)
    row_identifier = sku if sku else f"שורת Excel מספר {excel_row_number}"

    # בדיקת ערכים חסרים
    for column_name, display_name in REQUIRED_VALUE_COLUMNS.items():
        value = val(row, column_name)

        if not value:
            errors.append(
                f"מק״ט {row_identifier}: חסר ערך בעמודה '{display_name}'"
            )

    # בדיקת הברקוד
    barcode_original = val(row, COL_BARCODE)
    barcode_digits = clean_digits(barcode_original)

    if not barcode_original:
        # כבר תועדה שגיאת "חסר ברקוד"
        pass

    elif not barcode_original.replace(" ", "").replace("-", "").isdigit():
        errors.append(
            f"מק״ט {row_identifier}: הברקוד מכיל תווים שאינם ספרות "
            f"(הערך שהתקבל: {barcode_original})"
        )

    elif len(barcode_digits) != 13:
        errors.append(
            f"מק״ט {row_identifier}: הברקוד חייב להכיל בדיוק 13 ספרות "
            f"(התקבלו {len(barcode_digits)} ספרות: {barcode_digits})"
        )

    return errors
def write_validation_report(
    output_root: Path,
    errors: list[str]
) -> Path:
    """
    יוצרת קובץ טקסט עם כל השגיאות.
    UTF-8-SIG מאפשר פתיחה תקינה בעברית גם בפנקס הרשימות של Windows.
    """

    report_path = output_root / "שגיאות_נתונים.txt"

    lines = [
        "דוח שגיאות בקובץ האקסל",
        "=" * 40,
        "",
        f"סה״כ שגיאות: {len(errors)}",
        "",
    ]

    lines.extend(errors)

    report_path.write_text(
        "\n".join(lines),
        encoding="utf-8-sig"
    )

    return report_path
def safe_name(value: str) -> str:
    value = str(value).strip()
    value = re.sub(r'[\\/:*?"<>|]', "_", value)
    return value or "item"


def val(row, col, default=""):
    if col in row and pd.notna(row[col]):
        value = row[col]
        if isinstance(value, float) and value.is_integer():
            value = int(value)
        return str(value).strip()
    return default



def clean_digits(value: str) -> str:
    if value is None:
        return ""
    value = str(value).strip()
    if value.endswith(".0"):
        value = value[:-2]
    return re.sub(r"\D", "", value)


def sticker_dimension_mm(row, column_name: str, default_mm: float) -> float:
    """
    קורא מידה מהאקסל במילימטרים.
    תומך גם בקבצים ישנים שבהם נכתב 6/4 במקום 60/40.
    """
    raw = val(row, column_name, "")

    if not raw:
        return default_mm

    try:
        number = float(str(raw).replace(",", "."))
    except (TypeError, ValueError):
        return default_mm

    if number <= 0:
        return default_mm

    # תאימות לאקסלים ישנים שנכתבו בסנטימטרים: 6×4, 8×4 וכו'.
    if number <= 20:
        number *= 10

    return number




def copy_old_sku_folder(
    sku: str,
    product_dir: Path,
    source_roots: list[Path] | None = None,
) -> tuple[bool, str]:
    """
    מחפשת תיקייה ישנה בשם המק"ט בשלוש תיקיות מקור, לפי הסדר.

    אם נמצאה:
    - מעתיקה אותה לתוך תיקיית המוצר החדשה.
    - משנה את שם התיקייה המועתקת ל:
          <SKU> ישן
    """

    clean_sku = safe_name(sku)

    if not clean_sku:
        return False, "לא הוזן מק״ט"

    roots = source_roots or OLD_PRODUCTS_ROOTS
    source_folder = None

    raw_sku = str(sku).strip()
    possible_folder_names = []

    if raw_sku:
        possible_folder_names.append(raw_sku)

    if clean_sku not in possible_folder_names:
        possible_folder_names.append(clean_sku)

    for root in roots:
        if not root.exists() or not root.is_dir():
            continue

        for folder_name in possible_folder_names:
            candidate = root / folder_name

            if candidate.exists() and candidate.is_dir():
                source_folder = candidate
                break

        if source_folder is not None:
            break

    if source_folder is None:
        searched_paths = ", ".join(str(path) for path in roots)

        return (
            False,
            f"לא נמצאה תיקייה ישנה בשם המק״ט {sku}. "
            f"חיפוש בוצע בתיקיות: {searched_paths}",
        )

    destination_folder = product_dir / f"{clean_sku} ישן"

# בריצה שנייה: אם התיקייה כבר קיימת, מדלגים ולא מוחקים
    if destination_folder.exists():
        return (
            True,
            f"תיקיית היעד כבר קיימת, לא הועתקה מחדש: {destination_folder}"
    )

    shutil.copytree(
        source_folder,
        destination_folder,
)

    return True, str(destination_folder)


def normalize_certificate_number(value: str) -> str:
    """מנקה מספר תעודה שמגיע מאקסל."""
    if value is None:
        return ""

    value = str(value).strip()

    if value.endswith(".0"):
        value = value[:-2]

    return value


def copy_certificate_pdf(
    certificate_folder: Path,
    certificate_number: str,
    product_dir: Path,
    sku: str,
) -> tuple[bool, str]:
    """
    מחפשת PDF ששמו זהה למספר התעודה,
    מעתיקה אותו לתיקיית המק"ט ומשנה את שמו ל:
    EN <SKU>.pdf
    """

    certificate_number = normalize_certificate_number(
        certificate_number
    )

    if not certificate_number:
        return False, "לא הוזן מספר תעודה"

    if not certificate_folder.exists():
        return False, f"תיקיית התעודות לא קיימת: {certificate_folder}"

    candidates = [
        certificate_folder / f"{certificate_number}.pdf",
        certificate_folder / f"{certificate_number}.PDF",
    ]

    source_pdf = next(
        (path for path in candidates if path.exists()),
        None,
    )

    if source_pdf is None:
        target_stem = certificate_number.casefold()

        for path in certificate_folder.iterdir():
            if (
                path.is_file()
                and path.suffix.casefold() == ".pdf"
                and path.stem.casefold() == target_stem
            ):
                source_pdf = path
                break

    if source_pdf is None:
        return False, f"לא נמצא PDF עבור תעודה {certificate_number}"

    destination_pdf = product_dir / f"EN {safe_name(sku)}.pdf"
    shutil.copy2(source_pdf, destination_pdf)

    return True, str(destination_pdf)


def local_file_to_data_uri(path_value: str) -> str:
    if not path_value:
        return ""

    path = Path(str(path_value).strip().strip('"'))
    if not path.exists():
        return ""

    ext = path.suffix.lower().replace(".", "")
    if ext == "jpg":
        ext = "jpeg"
    if ext not in ["png", "jpeg", "webp", "gif"]:
        ext = "png"

    data = base64.b64encode(path.read_bytes()).decode("ascii")
    return f"data:image/{ext};base64,{data}"
def get_logo_data_uri():
    logo_path = Path(__file__).parent / "לוגו.jpeg"
    return local_file_to_data_uri(str(logo_path))

def barcode_url(barcode_value: str) -> str:
    """
    יוצר ברקוד EAN-13 באמצעות Zint,
    מחזיר אותו כ-Data URI מסוג SVG לשתילה בתוך HTML.
    """

    digits = clean_digits(barcode_value)

    if len(digits) != 13:
        raise ValueError(
            f"ברקוד EAN-13 חייב להכיל בדיוק 13 ספרות. התקבל: {digits}"
        )

    # מיקום zint.exe בתוך תיקיית הפרויקט
    zint_exe = Path(__file__).parent / "zint-2.16.0" / "zint.exe" 

    if not zint_exe.exists():
        raise FileNotFoundError(
            f"לא נמצא Zint בנתיב: {zint_exe}"
        )

    barcode_folder = Path(__file__).parent / "_barcodes"
    barcode_folder.mkdir(parents=True, exist_ok=True)

    svg_path = barcode_folder / f"{digits}_h45_g5_t1.svg"

    # אם הברקוד כבר נוצר בעבר, אין צורך לייצר שוב
    if not svg_path.exists():
        command = [
            str(zint_exe),

            # EAN-13 / EANX
            "--barcode=EANX",

            # הברקוד המלא
            f"--data={digits}",

            # קובץ SVG
            f"--output={svg_path}",

            # גובה הפסים
            "--height=39",

            # הורדת פסי השמירה מתחת לפסים הרגילים
            "--guarddescent=5",

            # רווח בין הפסים למספרים
            "--textgap=1",

            # אזורים לבנים תקניים בצדדים
            "--quietzones",

            # הטמעת הפונט בתוך SVG
            "--embedfont",
        ]

        result = subprocess.run(
            command,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
        )

        if result.returncode != 0:
            raise RuntimeError(
                "Zint נכשל ביצירת הברקוד:\n"
                + result.stderr
            )

        if not svg_path.exists():
            raise RuntimeError(
                f"Zint לא יצר את הקובץ: {svg_path}"
            )

    svg_data = svg_path.read_bytes()
    encoded = base64.b64encode(svg_data).decode("ascii")

    return f"data:image/svg+xml;base64,{encoded}"


def html_to_pdf(html: str, output_pdf: Path):
    output_pdf.parent.mkdir(parents=True, exist_ok=True)

    with sync_playwright() as p:
        browser = p.chromium.launch()

        page = browser.new_page(
            device_scale_factor=1
        )

        page.emulate_media(media="print")

        page.set_content(
            html,
            wait_until="networkidle"
        )

        # ממתינים לטעינת פונטים ותמונות
        page.evaluate("""
        async () => {
            if (document.fonts && document.fonts.ready) {
                await document.fonts.ready;
            }

            const images = Array.from(document.images);

            await Promise.all(
                images.map(async (img) => {
                    if (img.complete) return;

                    try {
                        await img.decode();
                    } catch (error) {
                        // ממשיכים גם אם תמונה מסוימת לא נטענה
                    }
                })
            );
        }
        """)

        # הקטנת טקסט אוטומטית עד שהוא נכנס לרוחב המוגדר
        page.evaluate("""
        () => {
            const MM_TO_PX = 96 / 25.4;
            const elements =
                document.querySelectorAll(".auto-fit-text");

            elements.forEach((element) => {
                const computed =
                    window.getComputedStyle(element);

                let fontSizePx =
                    parseFloat(computed.fontSize);

                const minimumMm =
                    parseFloat(
                        element.dataset.minFontMm || "1"
                    );

                const minimumPx =
                    minimumMm * MM_TO_PX;

                // מקטינים כל פעם בצעד קטן לקבלת התאמה מדויקת
                const stepPx = 0.15;

                while (
                    element.scrollWidth >
                        element.clientWidth + 0.25 &&
                    fontSizePx > minimumPx
                ) {
                    fontSizePx = Math.max(
                        minimumPx,
                        fontSizePx - stepPx
                    );

                    element.style.fontSize =
                        fontSizePx + "px";
                }
            });
        }
        """)

        page.pdf(
            path=str(output_pdf),
            print_background=True,
            prefer_css_page_size=True,
            scale=1,
        )

        browser.close()

        browser.close()
def extract_excel_images_by_sku(excel_path: Path, output_folder: Path):
    wb = load_workbook(excel_path)
    ws = wb.active

    output_folder.mkdir(parents=True, exist_ok=True)

    image_map = {}

    # עמודת המק"ט אצלך היא A
    sku_column = 1

    for img in ws._images:
        anchor = img.anchor._from
        excel_row = anchor.row + 1

        sku = ws.cell(row=excel_row, column=sku_column).value
        if not sku:
            continue

        sku = safe_name(str(sku).replace(".0", ""))

        img_path = output_folder / f"{sku}.jpeg"

        with open(img_path, "wb") as f:
            f.write(img._data())

        image_map[sku] = img_path

    return image_map


CARTON_HTML = """

<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<style>
@page {
    size: 210mm 160mm;
    margin: 0;
}

html,body{
    margin:0;
    padding:0;
    width:210mm;
    height:160mm;
    background:#fff;
    font-family:Arial,sans-serif;
    overflow:hidden;
}

.page{
    position:relative;
    width:210mm;
    height:160mm;
    box-sizing:border-box;
    padding:8mm;
}

/*************** TOP ***************/

.bsd{
    position:absolute;
    top:6mm;
    right:8mm;
    font-size:10mm;
    font-weight:bold;
    direction:rtl;
}

.product{
    position:absolute;
    top:11mm;
    left:35mm;
    width:60mm;
    height:50mm;
    object-fit:contain;
}

.logo{
    position:absolute;
    top:11mm;
    right:45mm;
    width:57mm;
    height:50mm;
    object-fit:contain;
}

/*************** DATA ***************/

.item-row{
    position:absolute;
    left:18mm;
    top:68mm;
    font-size:15mm;
}

.item-value{
    display:inline-block;
    width:92mm;
    border-bottom:0.6mm solid #000;
    text-align:center;
}

.pcs-row{
    position:absolute;
    left:18mm;
     top:92mm;
    font-size:15mm;
}

.pcs-value{
    display:inline-block;
    width:48mm;
    border-bottom:0.6mm solid #000;
    text-align:center;
}

.cbm-row{
    position:absolute;
    left:18mm;
    top:118mm;

    font-size:15mm;
}

.cbm-value{
    display:inline-block;
    width:62mm;
    border-bottom:0.6mm solid #000;
    text-align:center;
}

/*************** BARCODE ***************/
.barcode-box{
    position:absolute;

    right:8mm;
    bottom:14mm;

    width:84mm;
    height:54mm;

    border:0.6mm solid #000;

    box-sizing:border-box;
    text-align:center;
    direction:ltr;

    padding:2mm 3mm 2.5mm 3mm;
    overflow:hidden;
}
.barcode-title{
    font-size:7.2mm;
    line-height:1;
    margin-bottom:1.2mm;
    white-space:nowrap;
}

.barcode{
    width: {{ barcode_width_mr }}mm;
    max-width: 100%;
    height: auto;
    display: block;
    margin: 0 auto;
}

.barcode-sku{
    font-size:7mm;
    line-height:1;
    direction:rtl;
    margin-top:1.5mm;
    white-space:nowrap;
}
.sku-row{
    position:absolute;
    left:18mm;
    top:137mm;
    font-size:14mm;
    font-weight:400;
}

.sku-value{
    display:inline-block;
    width:70mm;
    text-align:center;
    border-bottom:0.6mm solid #000;
    margin-left:6mm;
}
</style>
</head>
<body>
<div class="page">

  <div class="top">
    <div class="bsd">בס"ד</div>

    {% if photo_data %}
      <img src="{{ photo_data }}" class="product">
    {% endif %}

   <img src="{{logo}}" class="logo">
  </div>

  <div class="item-row">
    ITEM NO:<span class="item-value">{{ item_no }}</span>
  </div>

  <div class="pcs-row">
    PCS:<span class="pcs-value">{{ pcs }}</span>
  </div>

  <div class="cbm-row">
    CBM:<span class="cbm-value">{{ cbm }}</span>
    
  </div>
  <div class="sku-row">
    {{ sku }}
</div>

  <div class="barcode-box">

    <div class="barcode-title">
        ITEM NO: {{ item_no }}
    </div>

    <img class="barcode" src="{{ barcode_url }}">

    <div class="barcode-sku">
        מק"ט {{ sku }}
    </div>

</div>



</div>
</body>
</html>
"""


BARCODES_84_HTML = """
<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<style>
  @page {
    size: A4 portrait;
    margin: 3.2mm;
  }

  html, body {
    margin: 0;
    padding: 0;
    background: #ffffff;
    font-family: Arial, Helvetica, sans-serif;
  }

  .page {
    width: 203.6mm;
    height: 290.6mm;
    display: grid;
    grid-template-columns: repeat(6, 1fr);
    grid-template-rows: repeat(14, 1fr);
    column-gap: 1.1mm;
    row-gap: 0.55mm;
    box-sizing: border-box;
    overflow: hidden;
  }

  .label {
    width: 100%;
    height: 100%;
    text-align: center;
    box-sizing: border-box;
    overflow: hidden;
    padding-top: 0.25mm;
  }

  .item{
    font-size:10px;
    font-weight:400;
    line-height:1.1;
    white-space:nowrap;
    margin-bottom:1mm;
  }

  .barcode{
    width: {{ barcode_width_hts }}mm;
    max-width: calc(100% - 1mm);
    height: auto;
    display: block;
    margin: 0 auto;
}

  .sku{
    font-size:9px;
    font-weight:400;
    line-height:1.1;
    direction:ltr;
    margin-top:0.5mm;
    white-space:nowrap;
    text-align:center;
  }
</style>
</head>

<body>
  <div class="page">
    {% for i in range(84) %}
      <div class="label">
        <div class="item">ITEM NO:{{ item_no }}</div>
        <img class="barcode" src="{{ barcode_url }}">
        <div class="sku">{{ sku }} מק&quot;ט</div>
      </div>
    {% endfor %}
  </div>
</body>
</html>
"""


WARNING_HTML = """
<!DOCTYPE html>
<html dir="rtl">
<head>
<meta charset="utf-8">
<style>
@page {
  size: {{width_warning}}mm {{height_warning}}mm;
  margin: 0;
}

html, body {
  width: {{width_warning}}mm;
  
  height: {{height_warning}}mm;
  margin: 0;
  padding: 0;
  font-family: Arial, Helvetica, sans-serif;
  direction: rtl;
  color: #000;
  background: #fff;
  overflow: hidden;
}

.wrapper {
  width: calc({{width_warning}}mm - 0.6mm);
  height: calc({{height_warning}}mm - 0.6mm);
  margin: 0.3mm;

  border: 0.35mm solid #000;
  box-sizing: border-box;
  display: flex;

  padding: 0.9mm 1mm;
  overflow: hidden;
  position: relative;
}

.right {
  width: 42%;
  padding-left: 0.8mm ;
  box-sizing: border-box;
  font-size: 1.55mm ;
  line-height: 1.35;
}

.left {
  width: 58%;
  box-sizing: border-box;
  font-size: 1.5mm;
  line-height: 1.14;
   padding-right: 1mm;
}

.title,
.warning-title,
.section-title {
  font-weight: bold;
  text-align: center;
  font-size:  1.6mm;
  margin: 0.25mm ;
}

.bold {
  font-weight: bold;
}

.barcode-box {
  border: 0.15mm  solid #000;
  margin-top: 0.7mm ;
  padding:
   0.45mm 
    0.4mm ;
  text-align: center;
  direction: ltr;
  box-sizing: border-box;
  overflow: hidden;
}

.item {
  font-size: 2.25mm ;
  line-height: 1;
  margin: 0 0 0.35mm  0;
  white-space: nowrap;
}
.barcode {
  width: min(26mm , 94%);
  height: auto;
  display: block;
  margin: 0 auto;
}

.sku {
  font-size: 2.1mm ;
  line-height: 1;
  direction: rtl;
  margin-top: 0.25mm ;
  white-space: nowrap;
}

ul {
  margin: 0;
  padding: 0;
  list-style: none;
}

li {
  position: relative;
  margin-bottom: 0.18mm;
  padding-right: 0.8mm;
  line-height: 1.14;
}

li::before {
  content: "•";
  position: absolute;
  right: 0;
  top: 0.03em;
  font-size: 1.30em;
  font-weight: bold;
  line-height: 1;
}
.producer {
  text-align: center;
  font-weight: bold;

  font-size: calc(1.6mm * {{ scale }});
  line-height: calc(1.05 * {{ scale_y }});

  margin-top: calc(0.3mm* {{ scale_y }});
  margin-bottom: calc(0.6mm * {{ scale_y }});

  white-space: nowrap;
}
.auto-fit-text {
  display: block;
  width: 100%;
  box-sizing: border-box;
  white-space: nowrap;
  overflow: hidden;
  text-overflow: clip;
}

</style>
</head>

<body>
<div class="wrapper">

  <div class="right">
    <div><span class="bold">יבואן:</span> שלום בן משה 26 בע"מ. ח.פ. 517009221</div>
    <div><span class="bold">כתובת:</span> כפר מסריק 
    <br/>
    <span class="bold">טלפון:</span> 077-2626260</div>

    <div class="warning-title">סחורה סוג א' תוצרת סין</div>

    <div>
      <span class="bold">אזהרת חנק:</span>
      המוצר עלול להכיל חלקים קטנים וחדים- ולכן עלול לגרום לחנק!
      הרחק ממקורות חום ואש!
    </div>
    {% if age =="0" %}
      <div><span class="bold">גיל מומלץ:</span>מוצר תינוקות מתאים לכל גיל</div>
    {% else %}
       
       <div><span class="bold">גיל מומלץ:</span> לא לשימוש ילדים מתחת גיל {{age}}.</div>
    {% endif %}

    

    <div class="producer">
      יצרן: {{producter}}
    </div>

    <div class="barcode-box">
      <div
        class="item auto-fit-text"
        data-min-font-mm="1.25"
>       ITEM NO: {{ item_no }}</div>
     <img class="barcode" src="{{ barcode_url }}">
      <div class="sku">מק&quot;ט {{ sku }}</div>
    </div>
  </div>

  <div class="left">
    <div class="title">אזהרות והוראות</div>
    <ul>
      <li>יש להסיר ולהרחיק את כל חומרי האריזה ושקיות הניילון מהישג ידם של ילדים ותינוקות לפני השימוש במוצר.</li>
      <li>השימוש במוצר בהשגחת מבוגר בלבד.</li>
      <li>לפני השימוש יש לבדוק את החיבורים, ברגים, רכיבים אלקטרוניים, סוללות ומטען במידה וקיימים במוצר.</li>
      <li>במידה ונמצא פגם, אין להשתמש במוצר טרם החלפתו.</li>
      <li>תכנים וצבעים עשויים להיות שונים מהמוצג באריזה.</li>
      <li>במוצרים המכילים צלילים: אין להצמיד לאוזן, עלול לגרום נזק לשמיעה.</li>
    </ul>

    <div class="section-title">שימוש בטוח בסוללות</div>
    <ul>
      <li>אין להטעין סוללות לא נטענות.</li>
      <li>יש להטעין סוללות תחת השגחת מבוגר.</li>
      <li>אין לערבב סוגים שונים של סוללות יחד או סוללות חדשות עם ישנות.</li>
      <li>יש להשתמש רק בסוללות מאותו הסוג כמומלץ.</li>
      <li>יש להכניס בטריות בקוטביות הנכונה (+/-).</li>
      <li>יש להשתמש בסוללות אלקליין בלבד.</li>
      <li>אין להשליך בטריות לאש, מפני סכנת פיצוץ.</li>
      <li>אין לחשוף את הסוללות למגע עם מים ורטיבות.</li>
      {% if  battery%}
          <li> 🔋 המשחק מופעל ע"י סוללות{battery}}</li>
  
      {% endif %}
    </ul>
    {{ warning_extra_html | safe }}
  </div>

</div>
</body>
</html>
"""
WARNING_EXTRA_HTML = {
    "1": "",  # רגיל - בלי תוספת

    "2": """
    
    <ul>
      <li>אין להטיס ליד אנשים, כבישים או חוטי חשמל.</li>
      <li>יש להשתמש ברחפן רק בהשגחת מבוגר.</li>
    </ul>
    """,

    "3": """
  
    <ul>
      <li>אין להשתמש בכביש או ליד רכבים נוסעים.</li>
      <li>אין להכניס אצבעות לגלגלים בזמן פעולה.</li>
    </ul>
    """
}


def render_templates(row, product_dir: Path, photo_data: str = ""):
    sku = val(row, COL_SKU)
    item_no = val(row, COL_ITEM_NO)
    pcs = val(row, COL_PCS)
    cbm = val(row, COL_CBM)
    barcode_value = val(row, COL_BARCODE)
    producter=val(row, COL_PRODUCER)
    age=val(row,COL_AGE)
    # המידות נלקחות מהאקסל; ברירת המחדל היא 60×40 מ"מ.
    # נתמך גם פורמט ישן של 6×4.
    x_html = sticker_dimension_mm(row, COL_X_STICKER, 80)
    Y_html = sticker_dimension_mm(row, COL_Y_STICKER, 40)
    barcode_digits = clean_digits(barcode_value)

    if len(barcode_digits) != 13:
        raise ValueError(
           f"ברקוד לא תקין למק״ט {sku}: {barcode_digits}"
        )

    barcode_first = barcode_digits[0]
    barcode_left = barcode_digits[1:7]
    barcode_right = barcode_digits[7:13]

    battery=val(row,COL_BATTERY)
    scale_x = x_html / 80
    scale_y = Y_html / 40    
    scale = (scale_x + scale_y) / 2
    warning_type = val(row, COL_WARNING_TYPE, "1")
    warning_extra_html = WARNING_EXTRA_HTML.get(warning_type, "")
    try:
        cbm = f"{float(cbm):.4f}"
    except (ValueError, TypeError):
        cbm = ""
    context = {
        "sku": sku,
        "item_no": item_no,
        "pcs": pcs,
        "cbm": cbm,
        "barcode": clean_digits(barcode_value),
        "barcode_url": barcode_url(barcode_value),
        "barcode_first": barcode_first,
        "barcode_left": barcode_left,
        "barcode_right": barcode_right,
        "photo_data": photo_data,
        "range": range,
        "logo": local_file_to_data_uri(str(LOGO_PATH)),
        "width_warning":x_html,
        "height_warning":Y_html,

        "battery":battery,
        "age": age,
        "producter": producter,
        "scale_x": scale_x,
        "scale_y": scale_y,
        "scale": scale,
        "warning_extra_html": warning_extra_html,
        "barcode_width_mr": BARCODE_WIDTH_MR_MM,
        "barcode_width_hts": BARCODE_WIDTH_HTS_MM,
        "barcode_width_st": BARCODE_WIDTH_ST_MM,
    }
    
   

    html_to_pdf(Template(CARTON_HTML).render(**context), product_dir / f"{safe_name(sku)}_MR.pdf")
    html_to_pdf(Template(BARCODES_84_HTML).render(**context), product_dir / f"{safe_name(sku)}_HTS.pdf")
    html_to_pdf(Template(WARNING_HTML).render(**context), product_dir / f"{safe_name(sku)}_ST.pdf")


def process_excel(
    excel_path: str,
    output_root: str,
    certificate_folder: str = "",
    progress_callback=None
):
    excel_path = Path(excel_path)
    output_root = Path(output_root)
    output_root.mkdir(parents=True, exist_ok=True)

    certificate_folder_path = (
        Path(certificate_folder)
        if certificate_folder
        else None
    )

    certificate_errors = []
    old_folder_errors = []

    # קריאת טבלת האקסל
    df = pd.read_excel(excel_path)
    df.columns = [str(column) for column in df.columns]

    # בדיקה שהעמודות עצמן קיימות בקובץ
    required_columns = [
        COL_SKU,
        COL_ITEM_NO,
        COL_PCS,
        COL_CBM,
        COL_BARCODE,
        COL_PHOTO,
        COL_WARNING_TYPE,
        COL_PRODUCER,
        COL_AGE,
        COL_X_STICKER,
        COL_Y_STICKER,
        COL_BATTERY,
        COL_CERTIFICATE,
    ]

    missing_columns = [
        column
        for column in required_columns
        if column not in df.columns
    ]

    if missing_columns:
        raise ValueError(
            "חסרות עמודות באקסל: "
            + ", ".join(missing_columns)
        )

    # ======================================
    # בדיקת תקינות של כל השורות לפני היצירה
    # ======================================

    all_errors = []
    valid_row_indexes = []

    for index, row in df.iterrows():
        # שורה 1 באקסל היא כותרת, לכן הנתונים מתחילים משורה 2
        excel_row_number = index + 2

        row_errors = validate_product_row(
            row,
            excel_row_number
        )

        if row_errors:
            all_errors.extend(row_errors)
        else:
            valid_row_indexes.append(index)

    # כתיבת דוח שגיאות
    report_path = None

    if all_errors:
        report_path = write_validation_report(
            output_root,
            all_errors
        )
    else:
        # אם נשאר דוח ישן מריצה קודמת – מוחקים אותו
        old_report = output_root / "שגיאות_נתונים.txt"

        if old_report.exists():
            old_report.unlink()

    # אם אין אף שורה תקינה, אין טעם להמשיך
    if not valid_row_indexes:
        raise ValueError(
            "לא נמצאו מוצרים תקינים ליצירת PDF.\n"
            f"דוח השגיאות נוצר כאן:\n{report_path}"
        )

    # ======================================
    # חילוץ תמונות רק לאחר בדיקת התקינות
    # ======================================

    extracted_images_dir = output_root / "_excel_images"

    image_map = extract_excel_images_by_sku(
        excel_path,
        extracted_images_dir
    )

    total = len(valid_row_indexes)

    # ======================================
    # יצירת PDF רק לשורות התקינות
    # ======================================

    for progress_index, dataframe_index in enumerate(
        valid_row_indexes,
        start=1
    ):
        row = df.loc[dataframe_index]

        sku = val(row, COL_SKU)
        item_no = val(row, COL_ITEM_NO)

        clean_sku = safe_name(sku)

        folder_name = (
            clean_sku
            or safe_name(item_no)
            or f"row_{dataframe_index + 2}"
        )

        product_dir = output_root / folder_name
        product_dir.mkdir(parents=True, exist_ok=True)

        # חיפוש והעתקת התיקייה הישנה של המק"ט.
        old_folder_copied, old_folder_result = copy_old_sku_folder(
            sku=sku,
            product_dir=product_dir,
        )

        if not old_folder_copied:
            old_folder_errors.append(
                f"מק״ט {sku}: {old_folder_result}"
            )

        # חיפוש התמונה שחולצה לפי המק"ט
        photo_path = image_map.get(clean_sku)

        photo_data = (
            local_file_to_data_uri(str(photo_path))
            if photo_path
            else ""
        )

        render_templates(
            row,
            product_dir,
            photo_data=photo_data
        )

        certificate_number = val(
            row,
            COL_CERTIFICATE,
        )

        if certificate_folder_path is not None:
            copied, certificate_result = copy_certificate_pdf(
                certificate_folder=certificate_folder_path,
                certificate_number=certificate_number,
                product_dir=product_dir,
                sku=sku,
            )
            if str(sku).strip() in {
                 "11798",
                 "17698",
                    "7928",
}:
                 print("=" * 60)
                 print("מק״ט:", repr(sku))
                 print("מספר תעודה שנקרא:", repr(certificate_number))
                 print("כל הערכים בשורה:")

                 for column_name, cell_value in row.items():
                    print(
                        repr(column_name),
                        "=>",
                        repr(cell_value)
                    )

            if not copied:
                certificate_errors.append(
                    f"מק״ט {sku}: {certificate_result}"
                )

        if progress_callback:
            progress_callback(progress_index, total)

    certificate_report_path = None

    if certificate_errors:
        certificate_report_path = (
            output_root / "שגיאות_תעודות.txt"
        )

        certificate_report_path.write_text(
            "\n".join(
                [
                    "דוח שגיאות בהעתקת תעודות",
                    "=" * 40,
                    "",
                    *certificate_errors,
                ]
            ),
            encoding="utf-8-sig",
        )

    old_folder_report_path = None

    if old_folder_errors:
        old_folder_report_path = (
            output_root / "שגיאות_תיקיות_ישנות.txt"
        )

        old_folder_report_path.write_text(
            "\n".join(
                [
                    "דוח שגיאות בהעתקת תיקיות ישנות",
                    "=" * 40,
                    "",
                    *old_folder_errors,
                ]
            ),
            encoding="utf-8-sig",
        )

    return {
        "created_products": total,
        "invalid_rows": len(df) - total,
        "error_count": len(all_errors),
        "report_path": report_path,
        "certificate_error_count": len(certificate_errors),
        "certificate_report_path": certificate_report_path,
        "old_folder_error_count": len(old_folder_errors),
        "old_folder_report_path": old_folder_report_path,
    }


